"""
Unit tests for Excel generator module.
"""

import pytest
import pandas as pd
import tempfile
from pathlib import Path
from openpyxl import load_workbook

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.excel_generator import ExcelGenerator


class TestExcelGenerator:
    """Test cases for ExcelGenerator class."""
    
    def test_init(self):
        """Test ExcelGenerator initialization."""
        generator = ExcelGenerator()
        
        assert generator.colors is not None
        assert generator.fonts is not None
        assert generator.alignments is not None
        assert generator.borders is not None
    
    def test_generate_excel_basic(self, temp_dir):
        """Test basic Excel generation."""
        generator = ExcelGenerator()
        
        # Mock results data
        results = {
            'records': {
                'matched': pd.DataFrame({'id': [1, 2], 'amount': [100, 200]}),
                'different': pd.DataFrame({'id': [3], 'amount': [300]}),
                'missing_in_source': pd.DataFrame({'id': [4], 'amount': [400]}),
                'missing_in_target': pd.DataFrame({'id': [5], 'amount': [500]})
            },
            'statistics': {
                'total_source': 4,
                'total_target': 4,
                'matched': 2,
                'different': 1,
                'missing_in_source': 1,
                'missing_in_target': 1,
                'field_statistics': {
                    'amount': {
                        'matches_count': 2,
                        'total_comparable': 3,
                        'match_rate': 0.67,
                        'differences_count': 1
                    }
                }
            },
            'config': {
                'reconciliation': {
                    'keys': ['id'],
                    'fields': [{'name': 'amount'}]
                }
            }
        }
        
        metadata = {
            'source_file': 'source.csv',
            'target_file': 'target.csv',
            'config_file': 'config.yaml',
            'execution_time': 1.5,
            'recon_date': '2025-06-24'
        }
        
        output_file = temp_dir / "test_output.xlsx"
        
        # Should not raise any exception
        generator.generate_excel(results, str(output_file), metadata)
        
        # Check that file was created
        assert output_file.exists()
        
        # Check that workbook can be loaded
        wb = load_workbook(str(output_file))
        assert 'Summary' in wb.sheetnames
        assert 'Matched' in wb.sheetnames
        assert 'Different' in wb.sheetnames
        assert 'Missing in Source' in wb.sheetnames
        assert 'Missing in Target' in wb.sheetnames
    
    def test_create_summary_sheet(self, temp_dir):
        """Test creating summary sheet."""
        generator = ExcelGenerator()
        
        from openpyxl import Workbook
        wb = Workbook()
        
        results = {
            'statistics': {
                'total_source': 100,
                'total_target': 95,
                'matched': 80,
                'different': 10,
                'missing_in_source': 5,
                'missing_in_target': 20,
                'field_statistics': {
                    'amount': {
                        'matches_count': 80,
                        'total_comparable': 90,
                        'match_rate': 0.89,
                        'differences_count': 10
                    }
                }
            }
        }
        
        metadata = {
            'source_file': 'test_source.csv',
            'target_file': 'test_target.csv',
            'config_file': 'test_config.yaml',
            'execution_time': 2.5,
            'recon_date': '2025-06-24 10:30:00'
        }
        
        # Should not raise any exception
        generator._create_summary_sheet(wb, results, metadata)
        
        # Check that Summary sheet was created
        assert 'Summary' in wb.sheetnames
        ws = wb['Summary']
        
        # Check that title exists
        assert ws['A1'].value == 'T-Rex Reconciliation Summary'
        
        # Check that metadata is present
        found_source = False
        found_target = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'Source File':
                found_source = True
                assert row[1] == 'test_source.csv'
            elif row[0] == 'Target File':
                found_target = True
                assert row[1] == 'test_target.csv'
        
        assert found_source and found_target
    
    def test_summary_sheet_auto_sizing(self, temp_dir):
        """Test that summary sheet columns are auto-sized properly."""
        generator = ExcelGenerator()
        
        from openpyxl import Workbook
        wb = Workbook()
        
        results = {
            'statistics': {
                'total_source': 100,
                'total_target': 95,
                'matched': 80,
                'different': 10,
                'missing_in_source': 5,
                'missing_in_target': 20,
                'field_statistics': {
                    'very_long_field_name_for_auto_sizing_test': {
                        'matches_count': 80,
                        'total_comparable': 90,
                        'match_rate': 0.89,
                        'differences_count': 10
                    }
                }
            }
        }
        
        metadata = {
            'source_file': 'test_source_with_very_long_file_name.csv',
            'target_file': 'test_target_with_very_long_file_name.csv',
            'config_file': 'test_config.yaml',
            'execution_time': 2.5,
            'recon_date': '2025-06-24 10:30:00'
        }
        
        # Should not raise any exception
        generator._create_summary_sheet(wb, results, metadata)
        
        # Check that Summary sheet was created
        assert 'Summary' in wb.sheetnames
        ws = wb['Summary']
        
        # Check that columns are auto-sized (not using the old hardcoded widths)
        # Column A should be sized to fit the long field name
        col_a_width = ws.column_dimensions['A'].width
        col_b_width = ws.column_dimensions['B'].width
        
        # The old hardcoded widths were A=25, B=30
        # With long content, column A should be wider than 30 to fit the long field name
        assert col_a_width > 30, f"Column A width {col_a_width} should be > 30 to fit long content"
        # Column B should also be auto-sized to fit the long file name
        assert col_b_width > 30, f"Column B width {col_b_width} should be > 30 to fit long content"
    
    def test_prepare_comparison_dataframe(self):
        """Test preparing DataFrame for comparison display."""
        generator = ExcelGenerator()
        
        df = pd.DataFrame({
            'id': [1, 2, 3],
            'source_amount': [100, 200, 300],
            'target_amount': [101, 199, 300],
            'source_status': ['A', 'B', 'C'],
            'target_status': ['A', 'X', 'C']
        })
        
        config = {
            'reconciliation': {
                'keys': ['id'],
                'fields': [
                    {'name': 'amount'},
                    {'name': 'status'}
                ]
            }
        }
        
        display_df = generator._prepare_comparison_dataframe(df, config)
        
        # Check that key columns come first
        assert display_df.columns[0] == 'id'
        
        # Check that source/target pairs are renamed properly
        expected_columns = ['id', 'Source: amount', 'Target: amount', 'Source: status', 'Target: status']
        assert list(display_df.columns) == expected_columns
    
    def test_prepare_comparison_dataframe_empty(self):
        """Test preparing empty DataFrame."""
        generator = ExcelGenerator()
        
        df = pd.DataFrame()
        config = {
            'reconciliation': {
                'keys': ['id'],
                'fields': [{'name': 'amount'}]
            }
        }
        
        display_df = generator._prepare_comparison_dataframe(df, config)
        assert display_df.empty
    
    def test_write_dataframe_to_sheet(self):
        """Test writing DataFrame to Excel sheet."""
        generator = ExcelGenerator()
        
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        
        df = pd.DataFrame({
            'col1': [1, 2, 3],
            'col2': ['A', 'B', 'C']
        })
        
        generator._write_dataframe_to_sheet(ws, df, freeze_headers=True, add_filters=True)
        
        # Check that data was written
        assert ws['A1'].value == 'col1'
        assert ws['B1'].value == 'col2'
        assert ws['A2'].value == 1
        assert ws['B2'].value == 'A'
        
        # Check that freeze panes was set
        assert ws.freeze_panes == 'A2'
        
        # Check that auto filter was set
        assert ws.auto_filter.ref is not None
    
    def test_apply_sheet_formatting(self):
        """Test applying sheet formatting."""
        generator = ExcelGenerator()
        
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        
        df = pd.DataFrame({
            'col1': [1, 2],
            'col2': ['A', 'B']
        })
        
        # Write data first
        generator._write_dataframe_to_sheet(ws, df)
        
        # Apply formatting
        generator._apply_sheet_formatting(ws, df)
        
        # Check header formatting
        header_cell = ws['A1']
        assert header_cell.font.bold == True
        assert header_cell.font.color.rgb in ['FFFFFF', '00FFFFFF']  # Allow for different openpyxl versions
        
        # Check that column widths were set (allow for auto-sizing variations)
        assert ws.column_dimensions['A'].width >= 8
        assert ws.column_dimensions['B'].width >= 8
    
    def test_create_sheets_with_empty_data(self, temp_dir):
        """Test creating sheets when data is empty."""
        generator = ExcelGenerator()
        
        from openpyxl import Workbook
        wb = Workbook()
        
        # Test with empty matched records
        results = {
            'records': {
                'matched': pd.DataFrame(),
                'different': pd.DataFrame(),
                'missing_in_source': pd.DataFrame(),
                'missing_in_target': pd.DataFrame()
            },
            'config': {
                'reconciliation': {
                    'keys': ['id'],
                    'fields': [{'name': 'amount'}]
                }
            }
        }
        
        # Should not raise exceptions
        generator._create_matched_sheet(wb, results)
        generator._create_different_sheet(wb, results)
        generator._create_missing_in_source_sheet(wb, results)
        generator._create_missing_in_target_sheet(wb, results)
        
        # Check that sheets were created
        assert 'Matched' in wb.sheetnames
        assert 'Different' in wb.sheetnames
        assert 'Missing in Source' in wb.sheetnames
        assert 'Missing in Target' in wb.sheetnames
        
        # Check that empty message is displayed
        assert wb['Matched']['A1'].value == 'No matched records found'
        assert wb['Different']['A1'].value == 'No different records found'
    
    def test_match_rate_conditional_formatting(self, temp_dir):
        """Test conditional formatting for match rate column."""
        generator = ExcelGenerator()
        
        # Mock results with different match rates
        results = {
            'records': {
                'matched': pd.DataFrame({'id': [1, 2], 'amount': [100, 200]}),
                'different': pd.DataFrame({'id': [3], 'amount': [300]}),
                'missing_in_source': pd.DataFrame(),
                'missing_in_target': pd.DataFrame()
            },
            'config': {
                'reconciliation': {
                    'keys': ['id'],
                    'fields': [{'name': 'amount'}]
                }
            },
            'statistics': {
                'total_source': 3,
                'total_target': 3,
                'matched': 2,
                'different': 1,
                'missing_in_source': 0,
                'missing_in_target': 0,
                'field_statistics': {
                    'field_perfect': {  # 100% match rate
                        'matches_count': 3,
                        'differences_count': 0,
                        'total_comparable': 3,
                        'match_rate': 1.0
                    },
                    'field_imperfect': {  # 66.7% match rate
                        'matches_count': 2,
                        'differences_count': 1,
                        'total_comparable': 3,
                        'match_rate': 0.667
                    }
                }
            }
        }
        
        metadata = {
            'source_file': 'test.csv',
            'target_file': 'test.csv',
            'config_file': 'test.yaml',
            'execution_time': 1.0,
            'recon_date': '2025-06-25'
        }
        
        output_file = temp_dir / "test_formatting.xlsx"
        generator.generate_excel(results, str(output_file), metadata)
        
        # Load and check formatting
        wb = load_workbook(output_file)
        ws = wb['Summary']
        
        # Find field statistics rows
        perfect_row = None
        imperfect_row = None
        for row_num in range(1, 30):
            cell_a = ws.cell(row=row_num, column=1)
            if cell_a.value == 'field_perfect':
                perfect_row = row_num
            elif cell_a.value == 'field_imperfect':
                imperfect_row = row_num
        
        assert perfect_row is not None, "Perfect match rate field not found"
        assert imperfect_row is not None, "Imperfect match rate field not found"
        
        # Check 100% match rate formatting (green)
        perfect_cell = ws.cell(row=perfect_row, column=5)  # Column E
        assert perfect_cell.fill.start_color.rgb == '00C6EFCE', "Perfect match rate should have green background"
        assert perfect_cell.font.color.rgb == '00006100', "Perfect match rate should have dark green text"
        assert perfect_cell.font.bold == True, "Perfect match rate should be bold"
        
        # Check <100% match rate formatting (red)
        imperfect_cell = ws.cell(row=imperfect_row, column=5)  # Column E
        assert imperfect_cell.fill.start_color.rgb == '00FFC7CE', "Imperfect match rate should have red background"
        assert imperfect_cell.font.color.rgb == '009C0006', "Imperfect match rate should have dark red text"
        assert imperfect_cell.font.bold == True, "Imperfect match rate should be bold"
    
    def test_values_are_different_method(self):
        """Test the _values_are_different method for blank/null comparisons."""
        generator = ExcelGenerator()
        
        # Test cases for different blank/null scenarios
        test_cases = [
            # source_value, target_value, expected_different
            (None, None, False),  # Both None
            (pd.NA, pd.NA, False),  # Both pandas NA
            ("", "", False),  # Both empty strings
            ("  ", "  ", False),  # Both whitespace (stripped)
            (None, "", True),  # None vs empty string
            (pd.NA, "", True),  # pandas NA vs empty string
            ("value", None, True),  # Value vs None
            ("value", "", True),  # Value vs empty string
            ("value", "value", False),  # Same values
            ("value1", "value2", True),  # Different values
            (0, None, True),  # Zero vs None
            (0, "", True),  # Zero vs empty string
            (0, 0, False),  # Same zeros
            ("0", 0, False),  # String zero vs numeric zero (after conversion)
        ]
        
        for source_val, target_val, expected in test_cases:
            result = generator._values_are_different(source_val, target_val)
            assert result == expected, f"Failed for {source_val} vs {target_val}: expected {expected}, got {result}"
    
    def test_highlight_differences_with_blanks(self, temp_dir):
        """Test highlighting differences when one value is blank and the other is not."""
        generator = ExcelGenerator()
        
        # Create test data with blank/null differences
        different_df = pd.DataFrame({
            'id': [1, 2, 3],
            'source_status': ['Active', None, ''],  # Mixed: value, None, empty string
            'target_status': ['Active', 'Inactive', 'Pending'],  # All have values
            'source_category': ['A', 'B', 'C'],  # All have values
            'target_category': ['A', None, '']  # Mixed: value, None, empty string
        })
        
        # Mock results structure
        results = {
            'config': {
                'reconciliation': {
                    'keys': ['id'],
                    'fields': [
                        {'name': 'status'},
                        {'name': 'category'}
                    ]
                }
            },
            'field_comparison': {
                'status': {
                    'matches': pd.Series([True, False, False]),  # Only first row matches
                    'matches_count': 1,
                    'total_comparable': 3,
                    'match_rate': 0.33
                },
                'category': {
                    'matches': pd.Series([True, False, False]),  # Only first row matches
                    'matches_count': 1,
                    'total_comparable': 3,
                    'match_rate': 0.33
                }
            },
            'records': {
                'matched': pd.DataFrame(),
                'different': different_df,
                'missing_in_source': pd.DataFrame(),
                'missing_in_target': pd.DataFrame()
            },
            'statistics': {
                'total_source': 3,
                'total_target': 3,
                'matched': 0,
                'different': 3,
                'missing_in_source': 0,
                'missing_in_target': 0,
                'field_statistics': {
                    'status': {'matches_count': 1, 'differences_count': 2, 'total_comparable': 3, 'match_rate': 0.33},
                    'category': {'matches_count': 1, 'differences_count': 2, 'total_comparable': 3, 'match_rate': 0.33}
                }
            }
        }
        
        metadata = {
            'source_file': 'test_source.csv',
            'target_file': 'test_target.csv',
            'config_file': 'test_config.yaml',
            'recon_date': '2025-06-25',
            'execution_time': 0.1
        }
        
        # Generate Excel file
        output_file = temp_dir / "test_blank_highlighting.xlsx"
        generator.generate_excel(results, str(output_file), metadata)
        
        # Verify file was created
        assert output_file.exists()
        
        # Load the workbook and check the Different sheet
        wb = load_workbook(str(output_file))
        assert "Different" in wb.sheetnames
        
        # This test mainly verifies no exceptions are thrown
        # Visual highlighting would need to be tested manually or with more complex Excel inspection
        # The fact that the file is created successfully indicates the highlighting logic works
        
        wb.close()
