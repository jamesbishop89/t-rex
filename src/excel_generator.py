"""
Excel Generator Module

Generates compre        # Fonts
        self.fonts = {
            'header': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'difference': Font(name='Calibri', size=10, bold=True, color='CC0000'),
            'normal': Font(name='Calibri', size=10),
            'title': Font(name='Calibri', size=12, bold=True),
            'summary_header': Font(name='Calibri', size=10, bold=True),
            'match_rate_perfect': Font(name='Calibri', size=10, bold=True, color='006400'),  # Dark green
            'match_rate_imperfect': Font(name='Calibri', size=10, bold=True, color='DC143C')  # Dark red
        }Excel output with multiple sheets for reconciliation results.
Includes formatting and detailed analysis.
"""

from pathlib import Path
from typing import Dict, List, Any, Tuple
import io

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.comments import Comment

from src.logger_setup import LoggerMixin


class ExcelGenerator(LoggerMixin):
    """
    Generator for comprehensive Excel reconciliation reports.
    
    Creates Excel files with multiple formatted sheets:
    - Summary: Overview with reconciliation and field statistics
    - Matched: Records that match within tolerances
    - Different: Records with differences highlighted
    - Missing in Target: Records only in source
    - Missing in Source: Records only in target
    """
    
    def __init__(self):
        """Initialize Excel generator with formatting styles."""
        self._setup_styles()
    
    def _setup_styles(self) -> None:
        """Set up Excel formatting styles for consistent appearance."""
        # Colors
        self.colors = {
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'difference_fill': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),
            'summary_header': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
            'match_rate_perfect': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'match_rate_imperfect': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        }
        
        # Fonts
        self.fonts = {
            'header': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'difference': Font(name='Calibri', size=11, bold=True, color='CC0000'),
            'normal': Font(name='Calibri', size=11),
            'summary_header': Font(name='Calibri', size=11, bold=True),
            'title': Font(name='Calibri', size=14, bold=True),
            'match_rate_perfect': Font(name='Calibri', size=11, bold=True, color='006100'),
            'match_rate_imperfect': Font(name='Calibri', size=11, bold=True, color='9C0006')
        }
        
        # Alignments
        self.alignments = {
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center')
        }
        
        # Borders
        thin_border = Side(border_style='thin', color='000000')
        self.borders = {
            'all': Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        }
    
    def generate_excel(self, results: Dict[str, Any], output_path: str, 
                      metadata: Dict[str, Any]) -> None:
        """
        Generate comprehensive Excel report with all reconciliation results.
        
        Args:
            results: Reconciliation results from ReconciliationEngine
            output_path: Path for output Excel file
            metadata: Execution metadata (file paths, timing, etc.)
            
        Raises:
            Exception: If Excel generation fails
        """
        self.logger.info(f"Generating Excel report: {output_path}")
        
        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Generate all sheets
            self._create_summary_sheet(wb, results, metadata)
            self._create_matched_sheet(wb, results)
            self._create_different_sheet(wb, results)
            self._create_missing_in_target_sheet(wb, results)
            self._create_missing_in_source_sheet(wb, results)
            
            # Save workbook
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            
            self.logger.info(f"Excel report generated successfully: {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, results: Dict[str, Any], 
                            metadata: Dict[str, Any]) -> None:
        """
        Create summary sheet with reconciliation and field statistics.
        
        Args:
            wb: Excel workbook
            results: Reconciliation results
            metadata: Execution metadata
        """
        self.logger.info("Creating Summary sheet")
        
        ws = wb.create_sheet("Summary", 0)
        stats = results['statistics']
        
        # Sheet title
        ws['A1'] = 'T-Rex Reconciliation Summary'
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:D1')
        
        # Metadata section
        metadata_data = [
            ['Source File', metadata.get('source_file', 'N/A')],
            ['Target File', metadata.get('target_file', 'N/A')],
            ['Config File', metadata.get('config_file', 'N/A')],
            ['Recon Date', metadata.get('recon_date', 'N/A')],
            ['Execution Time', f"{metadata.get('execution_time', 0):.2f} seconds"]
        ]
        
        start_row = 3
        for i, (label, value) in enumerate(metadata_data):
            row = start_row + i
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['summary_header']
        
        # Configuration section (moved above statistics)
        config_start_row = start_row + len(metadata_data) + 2
        ws[f'A{config_start_row}'] = 'Reconciliation Configuration'
        ws[f'A{config_start_row}'].font = self.fonts['title']
        
        # Get configuration details
        config = results['config']
        recon_config = config['reconciliation']
        source_columns = metadata.get('source_columns', [])
        
        # Order keys and fields according to source file column order
        ordered_keys = []
        ordered_fields = []
        
        # Create mapping for keys to check against source columns
        key_source_map = {}
        for k in recon_config['keys']:
            if isinstance(k, dict):
                source_col = k.get('source', k['name'])
                key_source_map[source_col] = k['name']
            else:
                key_source_map[k] = k
                
        # Create mapping for fields to check against source columns (exclude hidden fields)
        field_source_map = {}
        for f in recon_config['fields']:
            if f.get('hidden', False):
                continue
            source_col = f.get('source', f['name'])
            field_source_map[source_col] = f['name']
        
        for col in source_columns:
            if col in key_source_map:
                ordered_keys.append(key_source_map[col])
            
            if col in field_source_map and field_source_map[col] not in ordered_fields:
                ordered_fields.append(field_source_map[col])
        
        # Add any keys/fields that weren't found in source columns (fallback)
        for k in recon_config['keys']:
            name = k['name'] if isinstance(k, dict) else k
            if name not in ordered_keys:
                ordered_keys.append(name)
                
        for f in recon_config['fields']:
            if f.get('hidden', False):
                continue
            name = f['name']
            if name not in ordered_fields:
                ordered_fields.append(name)
        
        # Display reconciliation keys
        keys_row = config_start_row + 2
        ws[f'A{keys_row}'] = 'Reconciliation Keys'
        ws[f'B{keys_row}'] = ', '.join(ordered_keys) if ordered_keys else 'N/A'
        ws[f'A{keys_row}'].font = self.fonts['summary_header']
        
        # Display configured fields
        fields_row = keys_row + 1
        ws[f'A{fields_row}'] = 'Configured Fields'
        ws[f'B{fields_row}'] = ', '.join(ordered_fields) if ordered_fields else 'N/A'
        ws[f'A{fields_row}'].font = self.fonts['summary_header']
        
        # Display configured filters
        filters_row = fields_row + 1
        ws[f'A{filters_row}'] = 'Data Filters'
        
        filters_desc = []
        filters_config = recon_config.get('filters', {})
        
        for dataset_type in ['source', 'target']:
            if dataset_type in filters_config:
                for f in filters_config[dataset_type]:
                    desc = f"{dataset_type.title()}: {f['field']} {f['condition']}"
                    if 'value' in f:
                        desc += f" '{f['value']}'"
                    if 'values' in f:
                        desc += f" {f['values']}"
                    filters_desc.append(desc)
                    
        ws[f'B{filters_row}'] = '\n'.join(filters_desc) if filters_desc else 'None'
        ws[f'B{filters_row}'].alignment = Alignment(wrap_text=True, vertical='center')
        ws[f'A{filters_row}'].font = self.fonts['summary_header']
        
        # Statistics section (moved below configuration)
        stats_start_row = filters_row + 3
        ws[f'A{stats_start_row}'] = 'Reconciliation Statistics'
        ws[f'A{stats_start_row}'].font = self.fonts['title']
        
        stats_data = [
            ['Total Source Records', stats['total_source']],
            ['Total Target Records', stats['total_target']],
            ['Matched Records', stats['matched']],
            ['Different Records', stats['different']],
            ['Missing in Source', stats['missing_in_source']],
            ['Missing in Target', stats['missing_in_target']]
        ]
        
        for i, (label, value) in enumerate(stats_data):
            row = stats_start_row + 2 + i
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['summary_header']
        
          # Field statistics section if available
        if 'field_statistics' in stats and stats['field_statistics']:
            field_stats_start = stats_start_row + len(stats_data) + 3  # Adjusted for new layout
            ws[f'A{field_stats_start}'] = 'Field Statistics'
            ws[f'A{field_stats_start}'].font = self.fonts['title']
            
            # Headers for field statistics
            headers = ['Field', 'Matches', 'Differences', 'Total Comparable', 'Match Rate']
            for i, header in enumerate(headers):
                col = chr(ord('A') + i)
                ws[f'{col}{field_stats_start + 2}'] = header
                ws[f'{col}{field_stats_start + 2}'].font = self.fonts['summary_header']
                ws[f'{col}{field_stats_start + 2}'].fill = self.colors['summary_header']
            
            # Order field statistics by source file column order instead of by differences
            field_stats_dict = stats['field_statistics']
            ordered_field_stats = []
            
            # Add fields in source file order if they have statistics
            for col in source_columns:
                if col in field_stats_dict:
                    ordered_field_stats.append((col, field_stats_dict[col]))
            
            # Add any remaining fields that weren't in source columns (shouldn't happen normally)
            for field_name, field_stat in field_stats_dict.items():
                if field_name not in source_columns:
                    ordered_field_stats.append((field_name, field_stat))
              # Field statistics data
            for i, (field_name, field_stats) in enumerate(ordered_field_stats):
                row = field_stats_start + 3 + i
                ws[f'A{row}'] = field_name
                ws[f'B{row}'] = field_stats['matches_count']
                ws[f'C{row}'] = field_stats['differences_count']
                ws[f'D{row}'] = field_stats['total_comparable']
                # Set match rate as number with percentage format
                match_rate_cell = ws[f'E{row}']
                match_rate_cell.value = field_stats['match_rate']
                match_rate_cell.number_format = '0.00%'
                
                # Apply conditional formatting to match rate
                if field_stats['match_rate'] >= 1.0:  # 100% match rate
                    match_rate_cell.fill = self.colors['match_rate_perfect']
                    match_rate_cell.font = self.fonts['match_rate_perfect']
                else:  # Less than 100% match rate
                    match_rate_cell.fill = self.colors['match_rate_imperfect']
                    match_rate_cell.font = self.fonts['match_rate_imperfect']
        
        # Auto-size columns like other sheets
        self._auto_size_summary_columns(ws)
    
    def _create_matched_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Create sheet for matched records.
        
        Args:
            wb: Excel workbook
            results: Reconciliation results
        """
        self.logger.info("Creating Matched sheet")
        
        ws = wb.create_sheet("Matched")
        matched_df = results['records']['matched']
        
        if matched_df.empty:
            ws['A1'] = 'No matched records found'
            return
        
        # Prepare data for display
        display_df = self._prepare_comparison_dataframe(matched_df, results['config'])
        
        # Write data to worksheet
        self._write_dataframe_to_sheet(ws, display_df, freeze_headers=True, add_filters=True)
        
        # Apply formatting
        self._apply_sheet_formatting(ws, display_df)
        
        # Add transformation comments
        self._add_transformation_comments(ws, display_df, results)
        
        # Add post-merge adjustment comments (e.g. forced-zero MTM)
        self._add_post_merge_comments(ws, display_df, results)
    
    def _create_different_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Create sheet for different records with difference highlighting.
        
        Args:
            wb: Excel workbook
            results: Reconciliation results
        """
        self.logger.info("Creating Different sheet")
        
        ws = wb.create_sheet("Different")
        different_df = results['records']['different']
        
        if different_df.empty:
            ws['A1'] = 'No different records found'
            return
        
        # Prepare data for display
        display_df = self._prepare_comparison_dataframe(different_df, results['config'])
        
        # Write data to worksheet
        self._write_dataframe_to_sheet(ws, display_df, freeze_headers=True, add_filters=True)
        
        # Apply formatting and highlight differences
        self._apply_sheet_formatting(ws, display_df)
        self._highlight_differences(ws, different_df, results)
        
        # Add transformation comments
        self._add_transformation_comments(ws, display_df, results)
        
        # Add post-merge adjustment comments (e.g. forced-zero MTM)
        self._add_post_merge_comments(ws, display_df, results)
    
    def _create_missing_in_target_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Create sheet for records missing in target.
        
        Args:
            wb: Excel workbook
            results: Reconciliation results
        """
        self.logger.info("Creating Missing in Target sheet")
        
        ws = wb.create_sheet("Missing in Target")
        missing_df = results['records']['missing_in_target']
        
        if missing_df.empty:
            ws['A1'] = 'No records missing in target'
            return
        
        # Try to use raw data if available (using index tracking)
        use_raw = False
        if 'raw_data' in results and 'source' in results['raw_data'] and 'source___index__' in missing_df.columns:
            try:
                raw_source = results['raw_data']['source']
                indices = missing_df['source___index__']
                display_df = raw_source.loc[indices].copy()
                use_raw = True
                self.logger.info("Using raw data for Missing in Target sheet")
            except Exception as e:
                self.logger.warning(f"Failed to use raw data for Missing in Target: {e}")

        if not use_raw:
            # Get configured keys
            keys = results['config']['reconciliation']['keys']
            key_names = [k['name'] if isinstance(k, dict) else k for k in keys]
            fields = results['config']['reconciliation']['fields']
            hidden_field_names = {f['name'] for f in fields if f.get('hidden', False)}
            
            # Show keys and source columns (remove source_ prefix columns)
            source_columns = [col for col in missing_df.columns if col.startswith('source_') and col != 'source___index__'
                              and col.replace('source_', '') not in hidden_field_names]
            
            # Combine keys and source columns
            display_cols = []
            # Add keys first
            for key in key_names:
                if key in missing_df.columns:
                    display_cols.append(key)
            
            # Add source columns
            display_cols.extend(source_columns)
            
            display_df = missing_df[display_cols].copy()
            display_df.columns = [col.replace('source_', '') for col in display_df.columns]
        
        # Write data to worksheet
        self._write_dataframe_to_sheet(ws, display_df, freeze_headers=True, add_filters=True)
        
        # Apply formatting
        self._apply_sheet_formatting(ws, display_df)
    
    def _create_missing_in_source_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Create sheet for records missing in source.
        
        Args:
            wb: Excel workbook
            results: Reconciliation results
        """
        self.logger.info("Creating Missing in Source sheet")
        
        ws = wb.create_sheet("Missing in Source")
        missing_df = results['records']['missing_in_source']
        
        if missing_df.empty:
            ws['A1'] = 'No records missing in source'
            return
        
        # Try to use raw data if available (using index tracking)
        use_raw = False
        if 'raw_data' in results and 'target' in results['raw_data'] and 'target___index__' in missing_df.columns:
            try:
                raw_target = results['raw_data']['target']
                indices = missing_df['target___index__']
                display_df = raw_target.loc[indices].copy()
                use_raw = True
                self.logger.info("Using raw data for Missing in Source sheet")
            except Exception as e:
                self.logger.warning(f"Failed to use raw data for Missing in Source: {e}")

        if not use_raw:
            # Get configured keys
            keys = results['config']['reconciliation']['keys']
            key_names = [k['name'] if isinstance(k, dict) else k for k in keys]
            fields = results['config']['reconciliation']['fields']
            hidden_field_names = {f['name'] for f in fields if f.get('hidden', False)}
            
            # Show keys and target columns (remove target_ prefix columns)
            target_columns = [col for col in missing_df.columns if col.startswith('target_') and col != 'target___index__'
                              and col.replace('target_', '') not in hidden_field_names]
            
            # Combine keys and target columns
            display_cols = []
            # Add keys first
            for key in key_names:
                if key in missing_df.columns:
                    display_cols.append(key)
            
            # Add target columns
            display_cols.extend(target_columns)
            
            display_df = missing_df[display_cols].copy()
            display_df.columns = [col.replace('target_', '') for col in display_df.columns]
        
        # Write data to worksheet
        self._write_dataframe_to_sheet(ws, display_df, freeze_headers=True, add_filters=True)        
        # Apply formatting
        self._apply_sheet_formatting(ws, display_df)
    
    def _prepare_comparison_dataframe(self, df: pd.DataFrame, config: Dict[str, Any]) -> pd.DataFrame:
        """
        Prepare dataframe for comparison display with key columns first.
        
        Args:
            df: DataFrame to prepare
            config: Configuration dictionary
            
        Returns:
            pd.DataFrame: Prepared dataframe with organized columns
        """
        if df.empty:
            return df
        
        keys = config['reconciliation']['keys']
        fields = config['reconciliation']['fields']
        
        # Extract key names handling both string and dict formats
        key_names = [k['name'] if isinstance(k, dict) else k for k in keys]
        
        # Start with key columns
        display_columns = key_names.copy()
        
        # Track which fields get a diff % column
        diff_pct_fields = []
        
        # Add source/target pairs for each configured field (both comparison and ignored)
        for field in fields:
            # Skip hidden fields entirely from display
            if field.get('hidden', False):
                continue
            field_name = field['name']
            source_col = f'source_{field_name}'
            target_col = f'target_{field_name}'
            
            if source_col in df.columns and target_col in df.columns:
                display_columns.extend([source_col, target_col])
                
                # Add diff amount column (Source - Target) only if diff_amount: true
                if field.get('diff_amount', False):
                    source_numeric = pd.to_numeric(df[source_col], errors='coerce')
                    target_numeric = pd.to_numeric(df[target_col], errors='coerce')
                    
                    # Check if both values are numeric for at least some rows
                    if (source_numeric.notna().any() or target_numeric.notna().any()):
                        diff_amt_col = f'Diff Amount: {field_name}'
                        df = df.copy() if diff_amt_col not in df.columns else df
                        df[diff_amt_col] = source_numeric - target_numeric
                        display_columns.append(diff_amt_col)
                
                # Add diff % column only if field is flagged with diff_percent: true
                if field.get('diff_percent', False):
                    source_numeric = pd.to_numeric(df[source_col], errors='coerce')
                    target_numeric = pd.to_numeric(df[target_col], errors='coerce')
                    diff_col = f'diff_pct_{field_name}'
                    # Calculate diff %: (source - target) / target * 100, handle div-by-zero
                    with np.errstate(divide='ignore', invalid='ignore'):
                        diff_values = np.where(
                            target_numeric.abs() > 0,
                            ((source_numeric - target_numeric) / target_numeric.abs()) * 100,
                            np.where(source_numeric == target_numeric, 0.0, np.nan)
                        )
                    df = df.copy() if diff_col not in df.columns else df
                    df[diff_col] = np.round(diff_values, 4)
                    display_columns.append(diff_col)
                    diff_pct_fields.append(field_name)
        
        # Filter to available columns
        available_columns = [col for col in display_columns if col in df.columns]
        display_df = df[available_columns].copy()
          # Rename columns for better display, marking ignored fields
        column_renames = {}
        for col in display_df.columns:
            if col.startswith('source_'):
                field_name = col.replace('source_', '')
                # Check if this field is ignored
                field_config = next((f for f in fields if f['name'] == field_name), None)
                if field_config and field_config.get('ignore', False):
                    column_renames[col] = f"Source: {field_name} (ignored)"
                else:
                    column_renames[col] = f"Source: {field_name}"
            elif col.startswith('target_'):
                field_name = col.replace('target_', '')
                # Check if this field is ignored
                field_config = next((f for f in fields if f['name'] == field_name), None)
                if field_config and field_config.get('ignore', False):
                    column_renames[col] = f"Target: {field_name} (ignored)"
                else:
                    column_renames[col] = f"Target: {field_name}"
            elif col.startswith('diff_pct_'):
                field_name = col.replace('diff_pct_', '')
                column_renames[col] = f"Diff %: {field_name}"
        
        display_df = display_df.rename(columns=column_renames)
        
        return display_df
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, freeze_headers: bool = True,
                                add_filters: bool = True, auto_size: bool = True) -> None:
        """
        Write DataFrame to Excel worksheet with formatting options.
        
        Args:
            ws: Excel worksheet
            df: DataFrame to write
            freeze_headers: Whether to freeze header row
            add_filters: Whether to add auto filters
            auto_size: Whether to auto-size columns based on content
        """
        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Freeze headers if requested
        if freeze_headers and len(df) > 0:
            ws.freeze_panes = 'A2'
              # Add filters if requested
        if add_filters and len(df) > 0 and len(df.columns) > 0:
            # Create proper Excel column reference for any number of columns
            last_col = self._get_excel_column_name(len(df.columns))
            last_row = len(df) + 1
            filter_range = f"A1:{last_col}{last_row}"
            ws.auto_filter.ref = filter_range
            
        # Auto-size columns if requested
        if auto_size:
            self._auto_size_columns(ws, df)
    
    def _apply_sheet_formatting(self, ws, df: pd.DataFrame) -> None:
        """
        Apply consistent formatting to worksheet.
        
        Args:
            ws: Excel worksheet
            df: DataFrame (for determining dimensions)
        """
        if df.empty:
            return
        
        # Format headers
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.fonts['header']
            cell.fill = self.colors['header_fill']
            cell.alignment = self.alignments['center']
            cell.border = self.borders['all']
        
        # Format data cells
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = self.fonts['normal']
                cell.border = self.borders['all']        # Auto-size columns based on content
        self._auto_size_columns(ws, df)
    
    def _highlight_differences(self, ws, different_df: pd.DataFrame, results: Dict[str, Any]) -> None:
        """
        Highlight cells with differences in the Different sheet.
        Only highlights cells for fields that actually differ in each specific row.
        
        Args:
            ws: Excel worksheet
            different_df: DataFrame with different records
            results: Reconciliation results including field comparisons
        """
        if different_df.empty:
            return
        
        try:
            config = results['config']
            field_comparison = results.get('field_comparison', {})
            
            # Get column mapping
            display_df = self._prepare_comparison_dataframe(different_df, config)
            
            # For each row in different_df, determine which specific fields differ
            for diff_row_idx in range(len(different_df)):
                excel_row = diff_row_idx + 2  # +2 for header and 0-based index
                row_index = different_df.index[diff_row_idx]
                
                # Check each field to see if it differs in this specific row
                for field_name, comparison_result in field_comparison.items():
                    source_col_name = f"Source: {field_name}"
                    target_col_name = f"Target: {field_name}"
                    
                    if source_col_name not in display_df.columns or target_col_name not in display_df.columns:
                        continue
                    
                    # Determine if this field is a mismatch for this row using engine results
                    is_mismatch = False
                    if 'matches' in comparison_result and row_index in comparison_result['matches'].index:
                        # If it's False in matches, it's a mismatch
                        is_mismatch = not comparison_result['matches'].loc[row_index]
                    else:
                        # Fallback to naive comparison if engine result not found
                        source_value = different_df.iloc[diff_row_idx][f'source_{field_name}'] if f'source_{field_name}' in different_df.columns else None
                        target_value = different_df.iloc[diff_row_idx][f'target_{field_name}'] if f'target_{field_name}' in different_df.columns else None
                        is_mismatch = self._values_are_different(source_value, target_value)
                    
                    # Only highlight if this specific field differs in this specific row
                    if is_mismatch:
                        # Find column indices
                        source_col_idx = list(display_df.columns).index(source_col_name) + 1
                        target_col_idx = list(display_df.columns).index(target_col_name) + 1
                        
                        # Highlight source cell
                        source_cell = ws.cell(row=excel_row, column=source_col_idx)
                        source_cell.fill = self.colors['difference_fill']
                        source_cell.font = self.fonts['difference']
                        
                        # Highlight target cell
                        target_cell = ws.cell(row=excel_row, column=target_col_idx)
                        target_cell.fill = self.colors['difference_fill']
                        target_cell.font = self.fonts['difference']
        
        except Exception as e:
            self.logger.warning(f"Could not highlight differences: {e}")
    
    def _values_are_different(self, source_value, target_value) -> bool:
        """
        Check if two values are different, handling null, blank, and empty string comparisons.
        
        Args:
            source_value: Value from source dataset
            target_value: Value from target dataset
            
        Returns:
            bool: True if values are different, False if they are considered the same
        """
        import pandas as pd
        
        # Handle pandas NaN values
        source_is_na = pd.isna(source_value)
        target_is_na = pd.isna(target_value)
        
        # Both are NaN/None - considered same
        if source_is_na and target_is_na:
            return False
        
        # One is NaN/None, other is not - different
        if source_is_na != target_is_na:
            return True
        
        # Convert to string for comparison (handles various data types)
        source_str = str(source_value).strip() if source_value is not None else ""
        target_str = str(target_value).strip() if target_value is not None else ""
        
        # Compare string representations
        return source_str != target_str
    
    def _get_excel_column_name(self, col_num: int) -> str:
        """
        Convert column number to Excel column name (A, B, ..., Z, AA, AB, etc.).
        
        Args:
            col_num: Column number (1-based)
            
        Returns:
            Excel column name
        """
        result = ""
        while col_num > 0:
            col_num -= 1  # Convert to 0-based
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def _auto_size_columns(self, ws, df: pd.DataFrame, min_width: int = 8, max_width: int = 50) -> None:
        """
        Auto-size columns based on content length for better visibility.
        
        Args:
            ws: Excel worksheet
            df: DataFrame to analyze for column sizing
            min_width: Minimum column width
            max_width: Maximum column width
        """
        if df.empty:
            return
            
        for col_num in range(1, len(df.columns) + 1):
            column_letter = self._get_excel_column_name(col_num)
            
            # Get column name and calculate width needed for header
            header_name = df.columns[col_num - 1]
            header_width = len(str(header_name)) + 2  # Add padding
            
            # Calculate width needed for data in this column
            max_data_width = 0
            if len(df) > 0:
                for value in df.iloc[:, col_num - 1]:
                    if pd.notna(value):
                        value_width = len(str(value))
                        max_data_width = max(max_data_width, value_width)
            
            # Use the larger of header or data width, within limits
            optimal_width = max(header_width, max_data_width + 2)  # Add padding
            final_width = max(min_width, min(optimal_width, max_width))
            
            ws.column_dimensions[column_letter].width = final_width
    
    def _auto_size_summary_columns(self, ws, min_width: int = 8, max_width: int = 50) -> None:
        """
        Auto-size columns in summary sheet based on content length.
        
        Args:
            ws: Excel worksheet
            min_width: Minimum column width
            max_width: Maximum column width
        """
        # Get the maximum column index that contains data
        max_col = 1
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_col = max(max_col, cell.column)
        
        # Auto-size each column based on content
        for col_num in range(1, max_col + 1):
            column_letter = self._get_excel_column_name(col_num)
            max_width_needed = 0
            
            # Check all cells in this column to find the longest content
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value is not None:
                        cell_width = len(str(cell.value)) + 2  # Add padding
                        max_width_needed = max(max_width_needed, cell_width)
            
            # Set column width within the specified limits
            final_width = max(min_width, min(max_width_needed, max_width))
            ws.column_dimensions[column_letter].width = final_width
    
    def _add_transformation_comments(self, ws, display_df: pd.DataFrame, results: Dict[str, Any]) -> None:
        """
        Add Excel comments to cells showing transformation history (before/after values).
        
        Args:
            ws: Excel worksheet
            display_df: DataFrame being displayed in the sheet
            results: Reconciliation results containing transformation information
        """
        try:
            transformations = results.get('transformations', {})
            if not transformations:
                return
            
            # Get column mapping for display - handle renamed columns
            column_positions = {col: idx + 1 for idx, col in enumerate(display_df.columns)}
            
            # Get the original dataframe that contains the data before display preparation
            # We need to work with the raw merged data to access original column names
            sheet_name = ws.title
            if sheet_name == 'Matched':
                original_df = results['records']['matched']
            elif sheet_name == 'Different':
                original_df = results['records']['different']
            else:
                return  # Only add comments to matched and different sheets
            
            # Process transformations for both source and target
            for dataset_type in ['source', 'target']:
                if dataset_type not in transformations:
                    continue
                    
                for field_name, field_transformations in transformations[dataset_type].items():
                    # Find the display column name (renamed)
                    if dataset_type == 'source':
                        display_col_names = [
                            f'Source: {field_name}',
                            f'Source: {field_name} (ignored)'
                        ]
                        original_col_name = f'source_{field_name}'
                    else:  # target
                        display_col_names = [
                            f'Target: {field_name}',
                            f'Target: {field_name} (ignored)'
                        ]
                        original_col_name = f'target_{field_name}'
                    
                    # Find the display column that exists
                    display_col_name = None
                    for col_name in display_col_names:
                        if col_name in column_positions:
                            display_col_name = col_name
                            break
                    
                    if not display_col_name or original_col_name not in original_df.columns:
                        continue
                        
                    col_idx = column_positions[display_col_name]
                    
                    # Create mapping of transformation values to display rows
                    for transformation in field_transformations:
                        original_value = transformation['original_value']
                        new_value = transformation['new_value']
                        
                        # Look for the new_value in the original DataFrame column
                        value_matches = original_df[original_col_name] == new_value
                        
                        # Add comments to matching rows in the display
                        for idx, match in enumerate(value_matches):
                            if match:
                                excel_row = idx + 2  # +1 for header, +1 for 1-based indexing
                                
                                # Only proceed if this row exists in our display
                                if idx < len(display_df):
                                    # Build comment text
                                    comment_lines = ["🔄 Data Transformation Applied:"]
                                    if transformation['transformation_type'] == 'mapping':
                                        # For mappings, only show the mapping rule
                                        comment_lines.append(f"• Mapping: {transformation.get('mapping_rule', 'N/A')}")
                                    elif transformation['transformation_type'] == 'transformation':
                                        comment_lines.append(f"• Formula: {transformation.get('transformation_rule', 'N/A')}")
                                        comment_lines.append(f"• Before: {original_value}")
                                        comment_lines.append(f"• After: {new_value}")
                                    elif transformation['transformation_type'] == 'conditional':
                                        comment_lines.append(f"• Conditional: {transformation.get('condition_desc', 'N/A')}")
                                        comment_lines.append(f"• Before: {original_value}")
                                        comment_lines.append(f"• After: {new_value}")
                                    
                                    comment_text = "\n".join(comment_lines)
                                    
                                    # Create and add comment only if the cell has data
                                    cell = ws.cell(row=excel_row, column=col_idx)
                                    if cell.value is not None:
                                        # Check if comment already exists to avoid duplicates
                                        if not cell.comment:
                                            comment = Comment(comment_text, author="T-Rex Reconciliation")
                                            comment.width = 300
                                            comment.height = 150
                                            cell.comment = comment
                                            
                                            # Add a small indicator to show the cell has a transformation
                                            current_font = cell.font or self.fonts['normal']
                                            cell.font = Font(
                                                name=current_font.name,
                                                size=current_font.size,
                                                bold=current_font.bold,
                                                italic=True,  # Make italic to indicate transformation
                                                color=current_font.color
                                            )
                            
        except Exception as e:
            self.logger.warning(f"Failed to add transformation comments: {e}")

    def _add_post_merge_comments(self, ws, display_df: pd.DataFrame, results: Dict[str, Any]) -> None:
        """
        Add Excel comments to cells where post-merge calculations changed values
        (e.g. MTM forced to zero when settlement date equals reporting date).
        
        Args:
            ws: Excel worksheet
            display_df: DataFrame being displayed in the sheet
            results: Reconciliation results containing post-merge adjustment info
        """
        try:
            adjustments = results.get('post_merge_adjustments', {})
            if not adjustments:
                return
            
            column_positions = {col: idx + 1 for idx, col in enumerate(display_df.columns)}
            
            for key, adj_list in adjustments.items():
                # key format is "source_FIELDNAME" or "target_FIELDNAME"
                parts = key.split('_', 1)
                if len(parts) != 2:
                    continue
                dataset_type, field_name = parts
                
                # Find the display column
                display_col_names = [
                    f'Source: {field_name}' if dataset_type == 'source' else f'Target: {field_name}',
                    f'Source: {field_name} (ignored)' if dataset_type == 'source' else f'Target: {field_name} (ignored)'
                ]
                
                display_col_name = None
                for col_name in display_col_names:
                    if col_name in column_positions:
                        display_col_name = col_name
                        break
                
                if not display_col_name:
                    continue
                
                col_idx = column_positions[display_col_name]
                
                for adj in adj_list:
                    row_pos = adj['row_position']
                    if row_pos < len(display_df):
                        excel_row = row_pos + 2  # +1 header, +1 for 1-based
                        cell = ws.cell(row=excel_row, column=col_idx)
                        
                        original = adj['original_value']
                        new_val = adj['new_value']
                        
                        comment_text = (
                            f"\u26a0\ufe0f Post-Merge Adjustment:\n"
                            f"\u2022 Original value: {original}\n"
                            f"\u2022 Forced to: {new_val}\n"
                            f"\u2022 Reason: Settlement date = reporting date and target MTM is 0"
                        )
                        
                        if cell.comment:
                            # Append to existing comment
                            cell.comment.text += f"\n\n{comment_text}"
                        else:
                            comment = Comment(comment_text, author="T-Rex Reconciliation")
                            comment.width = 350
                            comment.height = 150
                            cell.comment = comment
                        
                        # Style the cell italic to indicate adjustment
                        current_font = cell.font or self.fonts['normal']
                        cell.font = Font(
                            name=current_font.name,
                            size=current_font.size,
                            bold=current_font.bold,
                            italic=True,
                            color=current_font.color
                        )
                        
        except Exception as e:
            self.logger.warning(f"Failed to add post-merge adjustment comments: {e}")
